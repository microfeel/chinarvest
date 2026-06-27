#!/usr/bin/env python3
"""Batch generate product JSONs from the 谷雨合作社 spreadsheet + 阴山优麦."""
import openpyxl, os, json, re, sys
from PIL import Image
Image.MAX_IMAGE_PIXELS = None

SRC = "/home/menxin/tradeMVP/docs/products/谷雨谷物合作社06月16日最新货盘.xlsx"
OUT = "/home/menxin/tradeMVP/data/products"
os.makedirs(OUT, exist_ok=True)

wb = openpyxl.load_workbook(SRC)

def slugify(name):
    s = name.lower().replace("\n", " ").strip()
    s = re.sub(r'[\s/\\]+', '-', s)
    s = re.sub(r'[()（）、，。/]', '', s)
    s = re.sub(r'[^\x00-\x7F]+', '', s)
    s = s.strip('-').strip()
    while '--' in s:
        s = s.replace('--', '-')
    return s[:60] or f"product"

def en_name(name):
    """Translate common Chinese product names to English."""
    cn_en = {
        "世小二": "ShiXiaoEr", "米不凡": "MiBuFan", "汤小郎": "TangXiaoLang",
        "格吉诺": "GeJiNuo", "珍稻优选": "ZhenDao", "汇知稻": "HuiZhiDao",
        "团圆粮": "TuanYuan", "一米距离": "YiMi", "实磨": "ShiMo",
        "天赐福米": "TianCiFu", "本真年代": "BenZhen", "森娇": "SenJiao",
        "远嫂": "YuanSao", "安社长": "AnSheZhang", "金吉顺": "JinJiShun",
        "友膳坊": "YouShanFang", "正小爽": "ZhengXiaoShuang",
        "雪韵龙江": "XueYunLongJiang",
        "有机黄小米": "Organic Yellow Millet", "东北黄豆": "Northeast Soybeans",
        "东北糯米": "Northeast Glutinous Rice", "七色糙米": "Seven-Color Brown Rice",
        "有机小米": "Organic Millet", "有机绿豆": "Organic Mung Beans",
        "有机红小豆": "Organic Red Beans", "有机黑芝麻": "Organic Black Sesame",
        "有机白芝麻": "Organic White Sesame", "有机花生": "Organic Peanuts",
        "东北玉米糁": "Northeast Corn Grits", "玉米糁": "Corn Grits",
        "小米面杂粮粉": "Millet Flour Blend", "大黄米杂粮粉": "Yellow Rice Flour Blend",
        "玉米面": "Corn Flour", "新玉米面条": "Fresh Corn Noodles",
        "优质小麦粉": "Premium Wheat Flour",
        "大碴子": "Corn Grits (Large)", "带豆": "with Beans", "无豆": "without Beans",
        "有机芸豆玉米碴": "Organic Kidney Bean & Corn Grits",
        "五常大米": "Wuchang Rice", "有机五常大米": "Organic Wuchang Rice",
        "圆粒珍珠米": "Round Pearl Rice", "富硒五常大米": "Selenium-Rich Wuchang Rice",
        "冷面": "Cold Noodles", "辣白菜": "Kimchi",
        "酸辣去骨鸡爪": "Spicy Sour Boneless Chicken Feet",
        "椒麻大杂烩": "Sichuan Pepper Mix", "佤味鸡爪": "Wa-Style Chicken Feet",
        "川味辣卤鸡爪": "Sichuan Spicy Braised Chicken Feet",
        "白桦树汁": "Birch Sap", "秋木耳": "Autumn Wood Ear Mushrooms",
        "辣椒酱": "Chili Sauce", "牛肉辣椒酱": "Beef Chili Sauce",
        "东北酸菜": "Northeast Sauerkraut", "纯猪肉红肠": "Pure Pork Sausage",
        "猪肉鸡肉红肠": "Pork & Chicken Sausage", "烤鹅": "Roast Goose",
        "北海海鸭蛋": "Beihai Sea Duck Eggs",
        "糯九玉米": "Nuojiu Glutinous Corn", "黑糯玉米": "Black Glutinous Corn",
    }
    for cn, en in cn_en.items():
        if cn in name:
            return en
    return name.strip().replace("\n", " ")[:40]

def en_category(cat):
    return {"grains": "Rice & Grains", "corn": "Fresh Corn", "specialty": "Specialty Foods"}.get(cat, cat)

def en_origin(name):
    if "五常" in name: return "Heilongjiang, Wuchang"
    if any(k in name for k in ["绥化", "有机黄小米", "有机小米", "有机绿豆", "有机红小豆", "有机黑芝麻", "有机白芝麻", "有机花生", "东北黄豆", "东北糯米"]): 
        return "Heilongjiang, Suihua"
    if "哈尔滨" in name or "东北" in name: return "Heilongjiang"
    if any(k in name for k in ["伊春", "秋木耳", "白桦树汁"]): return "Heilongjiang, Yichun"
    if "海伦" in name: return "Heilongjiang, Hailun"
    if "佳木斯" in name or "汤原" in name or "汤小郎" in name: return "Heilongjiang, Jiamusi"
    if "沈阳" in name or "冷面" in name or "酸菜" in name: return "Liaoning, Shenyang"
    if "大庆" in name or "肇州" in name: return "Heilongjiang, Daqing"
    if "山东" in name or "菏泽" in name: return "Shandong, Heze"
    if "北海" in name: return "Guangxi, Beihai"
    if "内蒙古" in name or "乌兰察布" in name: return "Inner Mongolia, Ulanqab"
    return "China"

def short_desc(name, spec, cat):
    if "米" in name and "稻" in name: return f"Premium {en_name(name)} from Heilongjiang, China's finest rice region. Quality grain for wholesale."
    if "小米" in name: return f"Premium organic millet from Heilongjiang — rich in nutrients, easy to digest. Ideal for wholesale."
    if "玉米" in name and "糯" in name: return f"Fresh sweet glutinous corn from Heilongjiang — naturally sweet, vacuum-sealed for freshness."
    if "黄豆" in name: return f"Northeast soybeans — non-GMO, protein-rich, ideal for food processing and wholesale."
    if any(k in name for k in ["绿豆","红小豆","黑芝麻","白芝麻","花生"]):
        return f"Organic {en_name(name).lower()} from Heilongjiang — premium quality for wholesale distribution."
    if "燕麦" in name or "oat" in spec.lower(): return f"Low GI oatmeal from Inner Mongolia — diabetic-friendly, high fiber, instant preparation."
    if "鸡爪" in name: return f"Premium {en_name(name)} — ready-to-eat, vacuum-packed for export. Popular in Asian markets."
    if "红肠" in name: return f"Traditional northeast Chinese sausage — smoked, ready-to-eat. Wholesale bulk packaging."
    if "烤鹅" in name: return f"Whole roasted goose — traditional northeast Chinese delicacy. Ready-to-eat, vacuum-packed."
    if "桦树" in name: return f"Pure birch sap from the Greater Khingan Range — natural, zero additives. Premium health beverage."
    if "木耳" in name: return f"Dried autumn wood ear mushrooms from Heilongjiang forests — premium quality for wholesale."
    if "冷面" in name: return f"Traditional northeast Korean cold noodles — ready-to-cook. Popular across Asian markets."
    if "酸菜" in name: return f"Traditional northeast Chinese sauerkraut — naturally fermented. Wholesale bulk available."
    if "海鸭蛋" in name: return f"Preserved sea duck eggs from Beihai — rich, savory flavor. Premium preserved egg product."
    if "辣椒酱" in name: return f"Traditional northeast chili sauce — made with premium ingredients. Wholesale bulk."
    return f"Premium {en_name(name)} from China — quality product for wholesale and export."

def gen_description(name, spec, cat):
    origin = en_origin(name)
    en = en_name(name)
    return f"{en} is a premium Chinese agricultural product sourced from {origin}. Our products are carefully selected, quality-controlled, and packaged for export. Suitable for wholesale distribution, food service, and retail channels. Complete export documentation available."

# Main categories from spreadsheet
SHEET_CATS = {"米面油杂粮系列": "grains", "玉米系列": "corn", "山野河鲜系列": "specialty"}
generated = 0
skipped = 0

for sheet_name, cat in SHEET_CATS.items():
    ws = wb[sheet_name]
    for r in range(3, ws.max_row + 1):
        name = str(ws.cell(r, 1).value or "").strip()
        spec = str(ws.cell(r, 3).value or "").strip()
        price = str(ws.cell(r, 5).value or "").strip()
        desc_cn = str(ws.cell(r, 8).value or "").strip()
        
        if not name or name in ["产品名", "||"] or len(name) < 2:
            continue
        
        clean_name = name.replace("\n", " ").replace("（", "(").replace("）", ")").strip()
        
        # Generate unique slug
        slug_base = slugify(clean_name)
        slug = slug_base
        count = 1
        while os.path.exists(os.path.join(OUT, f"{slug}.json")):
            count += 1
            slug = f"{slug_base}-{count}"
        
        # Clean spec for display
        spec_clean = spec.replace("\n", " ").strip()
        if not spec_clean:
            spec_clean = "Standard packaging"
        
        # Short description
        short = short_desc(clean_name, spec, cat)
        full_desc = gen_description(clean_name, spec, cat)
        if desc_cn and len(desc_cn) > 50:
            # Use AI translation later — for now use generated
            pass
        
        product = {
            "id": slug,
            "slug": slug,
            "category": cat,
            "subCategory": cat,
            "brand": clean_name.split("\n")[0] if "\n" in clean_name else clean_name.split()[0] if clean_name else "",
            "supplier": "谷雨谷物合作社",
            "origin": en_origin(clean_name),
            "certifications": ["有机认证"] if "有机" in clean_name else [],
            "specs": {
                "weight": spec_clean,
                "shelfLife": "12个月",
                "packaging": "Standard"
            },
            "export": {
                "hasExportHistory": False,
                "destinations": []
            },
            "images": {
                "hero": "/images/products/placeholder.jpg",
                "gallery": []
            },
            "i18n": {
                "en": {
                    "name": f"{en_name(clean_name)} {spec_clean}",
                    "shortDescription": short,
                    "description": full_desc,
                    "keywords": [en_name(clean_name).lower(), "china wholesale", cat, "export"]
                },
                "id": {
                    "name": en_name(clean_name),
                    "shortDescription": f"Produk pertanian China premium dari {en_origin(clean_name)}.",
                    "description": f"Produk pertanian China premium dari {en_origin(clean_name)}. Cocok untuk distribusi grosir, food service, dan ritel.",
                    "keywords": [en_name(clean_name).lower(), "grosir china", cat]
                },
                "ar": {
                    "name": en_name(clean_name),
                    "shortDescription": f"منتج زراعي صيني فاخر من {en_origin(clean_name)}.",
                    "description": f"منتج زراعي صيني فاخر من {en_origin(clean_name)}. مناسب لتوزيع الجملة وقطاع الخدمات الغذائية والتجزئة.",
                    "keywords": [en_name(clean_name).lower(), "الجملة الصين", cat]
                }
            }
        }
        
        filepath = os.path.join(OUT, f"{slug}.json")
        if not os.path.exists(filepath):
            with open(filepath, 'w') as f:
                json.dump(product, f, indent=2, ensure_ascii=False)
            print(f"  ✅ {clean_name[:35]:35s} → {slug}.json")
            generated += 1
        else:
            skipped += 1

print(f"\n{'='*50}")
print(f"Generated: {generated} new product JSONs")
print(f"Skipped (existed): {skipped}")
print(f"Total now: {len([f for f in os.listdir(OUT) if f.endswith('.json')])}")
