#!/usr/bin/env python3
"""Parse 谷雨谷物合作社 货盘 and generate product JSONs for chinarvest."""
import openpyxl, os, json
from PIL import Image
Image.MAX_IMAGE_PIXELS = None

SRC = "/home/menxin/tradeMVP/docs/products/谷雨谷物合作社06月16日最新货盘.xlsx"
OUT = "/home/menxin/tradeMVP/data/products"
os.makedirs(OUT, exist_ok=True)

wb = openpyxl.load_workbook(SRC)

# Sheet names
SHEETS = {
    "米面油杂粮系列": "grains",
    "玉米系列": "corn", 
    "山野河鲜系列": "specialty",
}

for sheet_name, category in SHEETS.items():
    ws = wb[sheet_name]
    print(f"\n=== {sheet_name} (category: {category}) ===")
    
    for r in range(3, ws.max_row + 1):
        name = str(ws.cell(r, 1).value or "").strip()
        spec = str(ws.cell(r, 3).value or "").strip()
        code = str(ws.cell(r, 4).value or "").strip()
        price = str(ws.cell(r, 5).value or "").strip()
        
        if not name or name in ["产品名", "||"] or len(name) < 2:
            continue
        
        # Generate slug from name
        slug = name.lower().replace("\n", " ").strip()
        slug = slug.replace(" ", "-").replace("--", "-").replace("（", "-").replace("）", "")
        slug = slug.replace("/", "-").replace("\\", "-")
        # Remove non-ascii for slug
        import re
        slug = re.sub(r'[^\x00-\x7F]+', '', slug)[:60].strip("-")
        if not slug:
            slug = f"product-{r}"
        
        product_id = f"{slug}-{r}"
        
        # Existing products - skip
        existing = os.listdir(OUT)
        if any(slug in f for f in existing):
            print(f"  SKIP (exists): {name[:30]}")
            continue
        
        print(f"  {name[:30]:30s} | {spec[:25]:25s}")

print("\nDone.")
